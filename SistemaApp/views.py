from django.views.generic import ListView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from core.mixins import PermitsPositionMixin
from django.contrib.auth.models import User
from django.db.models import Q
from django.utils import timezone
from UsuarioApp.models import Profile
from FabricaApp.models import FormularioProyectoFabrica
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle, Font, PatternFill
import pandas as pd

# Create your views here.


class DescargarProyectosFabricaView(LoginRequiredMixin, PermitsPositionMixin, View):
    def get(self, request, *args, **kwargs):

        proyectos = FormularioProyectoFabrica.objects.all().values(
            "codigo_sir",
            "nombre_propuesta",
            "fecha_inicio",
            "problema",
            "objetivo",
            "metodologia",
            "docente_id__nombre",
            "sede_id__sede_nombre",
            "empresa",
        )

        titulos = [
            "Id de proyecto",
            "Título del Proyecto",
            "Fecha inicio",
            "Contexto del Problema",
            "Objetivo del Proyecto",
            "Metodología",
            "Docente líder",
            "Sede",
            "Empresa asociada",
        ]

        df = pd.DataFrame(list(proyectos))
        df.columns = titulos

        wb = Workbook()
        ws = wb.active
        ws.title = "Proyectos Fabrica"

        date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")
        wb.add_named_style(date_style)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            ws.append(row)
            for c_idx, cell_value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if titulos[c_idx - 1] == "Fecha inicio":
                    cell.style = date_style

        # Create a table
        tab = Table(displayName="ProyectosTable", ref=ws.dimensions)

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # Style the header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Generate the response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="proyectos_fabrica.xlsx"'
        )

        # Save workbook to response
        wb.save(response)

        return response


class HomeView(LoginRequiredMixin, ListView):
    model = User
    template_name = "pages/index.html"

    def get_queryset(self):
        last_connected_users = User.objects.filter(
            Q(last_login__isnull=False)
        ).order_by("-last_login")[:5]
        return last_connected_users

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Agrega los usuarios activos al contexto
        recent_activity_cutoff = timezone.now() - timezone.timedelta(minutes=2)
        active_users = Profile.objects.filter(
            last_activity__gte=recent_activity_cutoff
        ).values_list("user_FK_id", flat=True)
        context["active_users"] = active_users

        # Agrega el recuento
        proyecto_count_ficha = FormularioProyectoFabrica.objects.count()
        latest_proyecto_ficha = FormularioProyectoFabrica.objects.order_by(
            "-registration_date"
        ).first()
        latest_registration_date_ficha = (
            latest_proyecto_ficha.registration_date if latest_proyecto_ficha else None
        )
        latest_registration_user_ficha = (
            latest_proyecto_ficha.user_id.username
            if latest_proyecto_ficha and latest_proyecto_ficha.user_id
            else None
        )

        context["proyecto_count_ficha"] = proyecto_count_ficha
        context["latest_registration_date_ficha"] = latest_registration_date_ficha
        context["latest_registration_user_ficha"] = latest_registration_user_ficha

        return context
